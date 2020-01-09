from django.shortcuts import render, redirect, HttpResponse
from .models import Post, Comment, Post_history
from django.contrib.auth.decorators import login_required
from .forms import post_create_form, post_update_form, post_comment_form
from django.contrib.auth.models import User
from django.contrib import messages
from recruitment_task import settings
from django.core.mail import send_mail
from openpyxl import Workbook


def home(request):
    context = {"posts": Post.objects.all().order_by("-date_posted")}
    return render(request, "blog/home.html", context)


@login_required
def post_detail(request, **kwargs):
    form = post_comment_form()
    feed = Post.objects.filter(id=int(kwargs["pk"]))
    context = {
        "posts": feed,
        "comments": Comment.objects.filter(post=feed.first()),
        "form": form,
    }
    if request.method == "POST":
        r = Comment(
            comment=request.POST["comment"], author=request.user, post=feed.first()
        )
        r.save()

    else:
        form
    return render(request, "blog/post_detail.html", context)


@login_required
def post_create(request):
    if request.method == "POST":
        title = request.POST["title"]
        content = request.POST["content"]
        post = Post(title=title, content=content, author=request.user)
        post.save()

        subject = f"New Feed from {request.user.username}"
        from_email = settings.DEFAULT_FROM_EMAIL

        for user in User.objects.exclude(username=request.user.username):
            if user.profile.notify:
                following_qs = user.geek.follow.all()
                for request.user in following_qs:
                    to_email = [user.email]
                    contact_message = f"A new feed with title {title}"
                    send_mail(
                        subject,
                        contact_message,
                        from_email,
                        to_email,
                        fail_silently=False,
                    )

        return redirect("blog-home")
    else:
        form = post_create_form()
    return render(request, "blog/post_create.html", {"form": form})


@login_required
def post_update(request, **kwargs):
    post = Post.objects.get(id=int(kwargs["pk"]))
    if request.user == post.author:
        if request.method == "POST":
            previous_title = post.title
            previous_content = post.content
            previous_date_posted = post.date_posted

            ph = Post_history(
                title=previous_title,
                content=previous_content,
                date_posted=previous_date_posted,
                update_of=post,
            )
            ph.save()

            post.title = request.POST["title"]
            post.content = request.POST["content"]
            post.save()
            messages.success(
                request, f'Blog Post "{previous_title}" has been successfully Updated.'
            )
            return redirect("blog-home")
        else:
            form = post_update_form({"title": post.title, "content": post.content})
    else:
        return HttpResponse(
            "<h1>You are not authorised to update a blog written by others, If you have administrative access kindly user admin login page.</h1>"
        )
    return render(request, "blog/post_update.html", {"form": form})


@login_required
def post_delete(request, **kwargs):
    post = Post.objects.get(id=int(kwargs["pk"]))

    context = {"post": post}
    if request.user == post.author:
        if request.method == "POST":
            messages.warning(
                request, f'Blog Post "{post.title}" has been successfully removed.'
            )
            post.delete()
            return redirect("blog-home")
    else:
        return HttpResponse(
            "<h1>You are not authorised to delete a blog written by others, If you have administrative access kindly user admin login page.</h1>"
        )
    return render(request, "blog/post_confirm_delete.html", context)


@login_required
def user_posts(request, *args, **kwargs):
    user = User.objects.get(id=int(kwargs["pk"]))
    posts = Post.objects.filter(author=user)
    logged_in = request.user
    context = {"user": user, "posts": posts, "logged_in": logged_in}
    return render(request, "blog/user_posts_list.html", context)


@login_required
def all_users(request, **kwargs):
    users = User.objects.exclude(id=request.user.id)
    context = {"users": users}
    return render(request, "blog/all_users.html", context)


@login_required
def post_like(request, **kwargs):
    post = Post.objects.get(id=kwargs["pk"])
    qs = post.liked_by.all()
    p = request.user
    if p not in qs:
        post.liked_by.add(p)
    else:
        post.liked_by.remove(p)
    return redirect("blog-home")


@login_required
def personalised_feed(request, **kwargs):
    user = User.objects.get(id=kwargs["pk"])

    followed_by = user.geek.follow.all()
    x = []
    for blogger in followed_by:
        q = Post.objects.filter(author_id=blogger.id)
        for r in q:
            x.append(r)

    context = {"posts": x}
    return render(request, "blog/personalised_feed.html", context)


@login_required
def following_list(request):
    user = User.objects.get(id=request.user.id)
    followed_by = user.geek.follow.all()
    context = {"followed_by": followed_by}
    return render(request, "blog/following.html", context)


@login_required
def user_report(request):
    if request.user.is_staff:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Users_Report"
        user_data = User.objects.all()
        j = 2
        head = [
            "Username",
            "User Email ID",
            "Total no. of Posts",
            "No. of times Reported",
            "is_staff",
        ]
        i = 1
        for data in head:
            ws1.cell(row=1, column=i).value = str(data)
            i += 1
        for user in user_data:
            i = 1
            k = Post.objects.filter(author=user)
            p = 0

            for post in k:
                p = p + post.reported_by.count()
            particulars = [
                user.username,
                user.email,
                Post.objects.filter(author_id=user.id).count(),
                p,
                user.is_staff,
            ]
            for data in particulars:
                ws1.cell(row=j, column=i).value = str(data)
                i += 1
            j += 1

        wb.save("Users_Report.xlsx")
        wb.close()
        messages.success(request, f"Users Report successfully created")

    return redirect("blog-home")


@login_required
def post_history(request, **kwargs):
    post_history_of = Post.objects.get(id=kwargs["pk"])
    posts = Post_history.objects.filter(update_of=post_history_of)

    context = {"posts": posts.all()}
    return render(request, "blog/post_history.html", context)

